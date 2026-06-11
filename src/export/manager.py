"""ExportManager — exportar dados em xlsx/csv/json (SPEC-13)."""
from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import Literal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.memory.database import Database

log = logging.getLogger(__name__)

QUERY_BASE = """
SELECT
    i.id,
    i.fonte_trilha,
    i.data_pub,
    i.titulo,
    i.link_original,
    i.veiculo,
    i.uf,
    i.midia,
    i.valor,
    i.publico,
    i.cm,
    i.assunto,
    i.sentimento AS sentimento_ccomsex,
    COALESCE(a.sentimento_ia, i.sentimento, 'NA') AS sentimento_exibido,
    a.narrativa,
    a.enquadramento,
    a.risco,
    a.confianca,
    a.tipo_inferencia,
    a.evidencia,
    i.source_file,
    i.ingested_at
FROM items i
LEFT JOIN analise_ia a ON a.item_id = i.id
{WHERE_CLAUSE}
ORDER BY i.data_pub DESC, i.id DESC
"""


class XLSXExporter:
    """Exporta para Excel."""

    CABECALHOS = [
        "ID",
        "Trilha",
        "Data",
        "Título",
        "Link",
        "Veículo",
        "UF",
        "Mídia",
        "Valor (R$)",
        "Público",
        "CM",
        "Assunto",
        "Sentimento CCOMSEx",
        "Sentimento Exibido",
        "Narrativa (IA)",
        "Enquadramento (IA)",
        "Risco (IA)",
        "Confiança IA",
        "Tipo Inferência",
        "Evidência",
        "Arquivo Origem",
        "Ingestão UTC",
    ]

    def exportar(self, rows: list[dict], destino: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ATALAIA Export"

        # Cabeçalho
        for col, cab in enumerate(self.CABECALHOS, 1):
            cell = ws.cell(row=1, column=col, value=cab)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="0D1F1A")
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for r_idx, row in enumerate(rows, 2):
            ws.append(list(row.values()))
            sent_col = self.CABECALHOS.index("Sentimento Exibido") + 1
            sent = row.get("sentimento_exibido", "")
            color = {
                "POSITIVA": "2eff8a",
                "NEGATIVA": "ff3b3b",
                "NEUTRA": "ffd84a",
            }.get(sent)
            if color:
                ws.cell(row=r_idx, column=sent_col).font = Font(
                    color=color, bold=True
                )

        # Auto-fit
        for col in ws.columns:
            max_len = max(
                (len(str(c.value or "")) for c in col), default=10
            )
            ws.column_dimensions[col[0].column_letter].width = min(
                max_len + 2, 60
            )

        wb.save(destino)


class CSVExporter:
    """Exporta para CSV."""

    def exportar(self, rows: list[dict], destino: Path) -> None:
        if not rows:
            destino.write_text("", encoding="utf-8-sig")
            return
        with destino.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)


class JSONExporter:
    """Exporta para JSON."""

    def exportar(self, rows: list[dict], destino: Path) -> None:
        destino.write_text(
            json.dumps(rows, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )


class ExportManager:
    """Orquestra exportação em múltiplos formatos."""

    def __init__(self, db: Database) -> None:
        self._db = db

    def exportar(
        self,
        formato: Literal["xlsx", "csv", "json"],
        destino: Path,
        trilha: str = "ALL",
        desde: str | None = None,
        ate: str | None = None,
        sentimento: str | None = None,
    ) -> int:
        """Exporta dados. Retorna número de linhas exportadas."""
        rows = self._buscar(trilha, desde, ate, sentimento)
        destino.parent.mkdir(parents=True, exist_ok=True)

        if formato == "xlsx":
            XLSXExporter().exportar(rows, destino)
        elif formato == "csv":
            CSVExporter().exportar(rows, destino)
        elif formato == "json":
            JSONExporter().exportar(rows, destino)

        log.info(
            "export.concluido",
            extra={
                "dados": {
                    "formato": formato,
                    "n": len(rows),
                    "destino": str(destino),
                }
            },
        )
        return len(rows)

    def _buscar(
        self, trilha: str, desde: str | None, ate: str | None, sentimento: str | None
    ) -> list[dict]:
        condicoes = []
        params = []
        if trilha != "ALL":
            condicoes.append("i.fonte_trilha = ?")
            params.append(trilha)
        if desde:
            condicoes.append("i.data_pub >= ?")
            params.append(desde)
        if ate:
            condicoes.append("i.data_pub <= ?")
            params.append(ate)
        where = (
            ("WHERE " + " AND ".join(condicoes)) if condicoes else ""
        )

        rows = [
            dict(r)
            for r in self._db.executar(
                QUERY_BASE.format(WHERE_CLAUSE=where), tuple(params)
            ).fetchall()
        ]
        if sentimento:
            rows = [
                r
                for r in rows
                if r.get("sentimento_exibido") == sentimento
            ]
        return rows

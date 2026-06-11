"""Testes do ExportManager (SPEC-13)."""
import json
import tempfile
from pathlib import Path

import openpyxl
import pytest

from src.export.manager import ExportManager, XLSXExporter, CSVExporter, JSONExporter
from src.memory.database import Database, MigrationRunner


ROWS_SAMPLE = [
    {
        "id": 1, "fonte_trilha": "A", "data_pub": "2026-01-15", "titulo": "Notícia Positiva",
        "link_original": "https://example.com/1", "veiculo": "Jornal A", "uf": "DF",
        "midia": "Online", "valor": 1000.0, "publico": 50000.0, "cm": None,
        "assunto": "DEFESA", "sentimento_ccomsex": "POSITIVA", "sentimento_exibido": "POSITIVA",
        "narrativa": None, "enquadramento": None, "risco": None, "confianca": None,
        "tipo_inferencia": None, "evidencia": None, "source_file": "test.xls", "ingested_at": "2026-01-15",
    },
    {
        "id": 2, "fonte_trilha": "B", "data_pub": "2026-02-10", "titulo": "Análise Negativa",
        "link_original": "https://example.com/2", "veiculo": "Portal B", "uf": "SP",
        "midia": "Impresso", "valor": 500.0, "publico": 30000.0, "cm": None,
        "assunto": "MILITAR", "sentimento_ccomsex": "NEUTRA", "sentimento_exibido": "NEGATIVA",
        "narrativa": "Narrativa X", "enquadramento": None, "risco": "ALTO", "confianca": 0.85,
        "tipo_inferencia": "INTERPRETACAO", "evidencia": None, "source_file": None, "ingested_at": "2026-02-10",
    },
]


class TestXLSXExporter:
    def test_gera_arquivo_valido(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        XLSXExporter().exportar(ROWS_SAMPLE, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # 1 cabeçalho + 2 linhas = 3 linhas total
        assert ws.max_row == 3
        path.unlink()

    def test_cabecalho_correto(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        XLSXExporter().exportar(ROWS_SAMPLE, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Sentimento Exibido" in headers
        assert "ID" in headers
        path.unlink()

    def test_banco_vazio_so_cabecalho(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = Path(f.name)
        XLSXExporter().exportar([], path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Arquivo criado mesmo sem linhas
        assert ws.max_row >= 1
        path.unlink()


class TestCSVExporter:
    def test_gera_csv_com_bom(self):
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = Path(f.name)
        CSVExporter().exportar(ROWS_SAMPLE, path)
        content = path.read_bytes()
        # BOM UTF-8
        assert content.startswith(b"\xef\xbb\xbf")
        path.unlink()

    def test_conteudo_correto(self):
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = Path(f.name)
        CSVExporter().exportar(ROWS_SAMPLE, path)
        text = path.read_text(encoding="utf-8-sig")
        assert "Jornal A" in text
        assert "Portal B" in text
        path.unlink()

    def test_banco_vazio(self):
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = Path(f.name)
        CSVExporter().exportar([], path)
        # Vazio retorna string vazia (com ou sem BOM)
        content = path.read_bytes()
        # pode ter BOM UTF-8 ou ser vazio — ambos são válidos
        assert len(content) <= 3
        path.unlink()


class TestJSONExporter:
    def test_json_valido(self):
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            path = Path(f.name)
        JSONExporter().exportar(ROWS_SAMPLE, path)
        loaded = json.loads(path.read_text(encoding="utf-8"))
        assert len(loaded) == 2
        path.unlink()

    def test_json_banco_vazio(self):
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            path = Path(f.name)
        JSONExporter().exportar([], path)
        loaded = json.loads(path.read_text(encoding="utf-8"))
        assert loaded == []
        path.unlink()


@pytest.fixture
def db_temp():
    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as f:
        path = Path(f.name)
    db = Database.inicializar(path)
    MigrationRunner(db).aplicar_todas()
    yield db
    path.unlink(missing_ok=True)


class TestExportManager:
    def test_banco_vazio_retorna_zero(self, db_temp):
        manager = ExportManager(db_temp)
        with tempfile.TemporaryDirectory() as tmp:
            n = manager.exportar("json", Path(tmp) / "out.json")
        assert n == 0

    def test_todos_formatos(self, db_temp):
        manager = ExportManager(db_temp)
        with tempfile.TemporaryDirectory() as tmp:
            for fmt in ["xlsx", "csv", "json"]:
                n = manager.exportar(fmt, Path(tmp) / f"out.{fmt}")
                assert isinstance(n, int)
